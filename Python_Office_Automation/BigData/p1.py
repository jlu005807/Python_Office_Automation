import random
import openpyxl

sn12=('11','19','34')
sn34=('20','19','18')
sn56=('03','02','01')

excel_filename = '大学计算机.xlsx'

wb = openpyxl.Workbook()

shee1 = wb.active
shee1.title = '理论课平时成绩'
shee1.column_dimensions['A'].width = 10

row=1
for clo in sn12:
    for i in sn34:
        for j in sn56:
         for sn78 in range(1,31):
            sn78=str(sn78)

            sn78=sn78.zfill(2)

            sn=clo+i+j+sn78
            print(sn)
            shee1.cell(row=row, column=1, value=sn)
            shee1.cell(row=row, column=2).number_format ='0'
            shee1.cell(row=row, column=2, value=random.randint(0,100))
            row+=1

wb.save(excel_filename)
wb.close()
