import openpyxl
excel_file = '大学计算机.xlsx'

wb=openpyxl.load_workbook(excel_file)

ws1=wb['理论课平时成绩']

ws2=wb['实验课成绩']

ws3=wb['期末考试成绩']

ws4=wb.create_sheet(title='总成绩',index=3)

ws4.column_dimensions['A'].width=10
ws4.column_dimensions['B'].width=15
ws4.column_dimensions['C'].width=15
ws4.column_dimensions['D'].width=15
ws4.column_dimensions['E'].width=15

nrows_ws1=ws1.max_row
nrows_ws2=ws2.max_row
nrows_ws3=ws3.max_row

for i in range(1,nrows_ws1+1):
    ws4.cell(row=i,column=1,value=ws1.cell(row=i,column=1).value)
    ws4.cell(row=i,column=2).number_format='0'
    ws4.cell(row=i,column=2,value=ws1.cell(row=i,column=2).value)

    for j in range(1,nrows_ws2+1):
        if ws2.cell(j,1).value==ws4.cell(i,1).value:
            ws4.cell(row=i,column=3).number_format='0'
            ws4.cell(row=i,column=3,value=ws2.cell(j,2).value)
            break

    for j in range(1,nrows_ws3+1):
        if ws3.cell(j,1).value==ws4.cell(i,1).value:
            ws4.cell(row=i,column=4).number_format='0'
            ws4.cell(row=i,column=4,value=ws3.cell(j,2).value)
            break

    ws4.cell(row=i,column=5).number_format='0'
    ws4.cell(row=i,column=5,value=ws4.cell(i,2).value+ws4.cell(i,3).value+ws4.cell(i,4).value)

ws4.insert_rows(1)
ws4.cell(row=1,column=1,value='学号')
ws4.cell(row=1,column=2,value='理论课平时成绩')
ws4.cell(row=1,column=3,value='实验课成绩')
ws4.cell(row=1,column=4,value='期末考试成绩')
ws4.cell(row=1,column=5,value='总成绩')

wb.save(excel_file)
wb.close()