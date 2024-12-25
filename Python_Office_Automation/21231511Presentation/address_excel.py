import pandas as pd
from openpyxl import load_workbook

# 处理日期
# 打开文件
workbook = load_workbook(filename='日化.xlsx', data_only=True)
# 选择工作表
ws = workbook['销售订单表']
nrows = ws.max_row

# 遍历第二列单元格
for i in range(2, nrows + 1):
    # 处理单元格数据
    if isinstance(ws.cell(i, 2).value, str):
        print(ws.cell(i, 2).value)
        ws.cell(i, 2).value = ws.cell(i, 2).value.replace("#", "/")

# 保存文件
workbook.save(filename='日化1.xlsx')
# 关闭文件
workbook.close()

# 读取"销售订单表"到DataFrame
df_dd = pd.read_excel('日化1.xlsx', sheet_name='销售订单表')
print(df_dd.head(20))

# 读取"商品信息表"到DataFrame
df_xx = pd.read_excel('日化1.xlsx', sheet_name='商品信息表')
print(df_xx.head(20))

# 处理缺失值
print(df_dd)
df_dd.dropna(inplace=True)
print(df_dd)

# 根据商品编号，增加商品小类列
dict_xx = dict()
for index, row in df_xx.iterrows():
    print(row['商品编号'], row['商品小类'])
    dict_xx[row['商品编号']] = row['商品小类']
print(dict_xx)
df_dd["商品小类"] = df_dd["商品编号"].replace(dict_xx)
print(df_dd.head(20))

# 排序，按商品小类排序也可以
df_dd.sort_values(by=['所在地市', '商品编号'], inplace=True)
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
print(df_dd.head(200))

# 并查看分组后的信息，此操作是为了便于同学们理解
# 可以直接遍历df_dd输出到word
print(df_dd.groupby(['所在地市', '商品编号']).sum())
print((df_dd.groupby(['所在地市', '商品编号']).sum()).describe())
# 保存到文件，此操作是为了便于同学们理解，
# 可以直接遍历df_dd输出到word
df_dd.to_excel('日化2.xlsx', index=False)